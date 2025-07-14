from DotaApi import initialize_connections
engine = initialize_connections.initialize_engine()
# MYSQL QUERY: i literally have no clue but it doesnt run here unless its at the very top. 
# So this intializes a connection and somehow it works.

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import matplotlib 
matplotlib.use('Agg')   
import pandas as pd
import matplotlib.pyplot as plt
from DotaApi import database_handler
from DotaApi import api_handler
from DotaApi import calculations
from DotaApi import database_handler
from DotaApi import winrate

#This file imports all the functions from the other files and uses it's own excelsheet creator function to output the master.xlsx excel sheet

pd.options.display.max_columns = None
pd.options.display.max_colwidth = None #for some reason pandas truncates df.to_string() if colwidth has a cap
#pd.set_option("display.width", 0)  # THIS ALLOWS ALL COLUMNS TO BE DISPLAYED, WITHOUT COLUMNS WILL BE "/" AND NEW LINED


#VARIABLES
#example comment

steam_id = 171262902 #watson
steam_id = 405788540

position = "POSITION_1"
isOnMyTeam = True #this is only used in player_graphs and worksheet string. by default its true for player_graphs
"========================================================"
minute = 11 #MINUTE 11 BY DEFAULT. minute 11 is exactly 10:01
skip_interval = 25
number_of_matches_to_parse = 10 #accepts numbers 0-{skip_interval}, for numbers above it needs to be intervals of {skip_interval}
"========================================================"

# Hard coded excel sheet generator. Goal is to move this into Django website.
def make_all_excel_sheets(df_raw, df_player_calculations, dict_of_plts, steam_id, position, minute, number_of_matches_to_parse, isOnMyTeam=True): #just for ease of use, so i dont have to call every one seperately
    minute = int(minute)
    position = str(position)
    number_of_matches_to_parse = int(number_of_matches_to_parse)

    file_name = "raw_data" + ".xlsx"
    df_raw.to_excel(file_name)

    for key in dict_of_plts:
        fig = dict_of_plts[key]
        fig.savefig(f"{key}.png", format="png")
        plt.close(fig)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Stats: (average data of everyone in the game, in relation to you)"
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A3"] = "Graphs: (plotted data of only you)"
    ws["A5"] = f"Parsing last {number_of_matches_to_parse} matches of {steam_id}. \nPARAMETERS: position={position}, ally={isOnMyTeam}, taking stats at minute {(int(minute)-1)}."
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 50
    ws["A6"] = "The networthDifference stats are ordered by comparing the respective positions to themselves for indices 1-5 (so index 3 is YourPos3-TheirPos3). Indices 6-10 are comparing positions to their lane opposition (index 8 is YourPos3-TheirPos1). The graph compares you agaisnt your lane opponent"
    ws["A6"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A7"] = "The stats above kills/deaths are the kda ratio"
    ws["A8"] = "If levels/kda is empty/lacking data that means stratz api returned no playbackData. Try parsing your latest match on stratz (idk if this will work)"
    ws["A8"].alignment = Alignment(wrap_text=True, vertical="top") 

    ws["B1"] = df_player_calculations["networth_difference"].to_string()
    ws["B1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["B"].width = 92
    ws.add_image(Image("networthDifference.png"), "B2") #start adding images below the df string

    ws["C1"] = df_player_calculations["lastHitsAverage"].to_string()
    ws["C1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["C"].width = 92
    ws.add_image(Image("lastHitsPerMinuteSum.png"), "C2")

    ws["D1"] = df_player_calculations["deniesAverage"].to_string()
    ws["D1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["D"].width = 92
    ws.add_image(Image("deniesPerMinuteSum.png"), "D2") 
    
    ws["E1"] = df_player_calculations["levelAverage"].to_string()
    ws["E1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["E"].width = 92
    ws.add_image(Image("level.png"), "E2") 

    ws["F1"] = df_player_calculations["kdaAverage"].to_string()
    ws["F1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["F"].width = 92
    ws.add_image(Image("kills.png"), "F2") 
    ws.add_image(Image("deaths.png"), "G2") 
    
    
    wb.save("master.xlsx")

    print("Excel Sheets created!")


#SCRIPT
def main_script(steam_id=171262902, position="POSITION_1", isOnMyTeam=True, minute=11, skip_interval=10, number_of_matches_to_parse=10):
    # Api querying, extraction of json response
    df_raw = api_handler.queries_to_batches_main(steam_id, position, skip_interval, number_of_matches_to_parse)
    #print(df_raw)
    #print("=================")

    # Constructon of calculated dataframes
    df_calculated = calculations.adding_columns(df_raw, steam_id, minute, isOnMyTeam, number_of_matches_to_parse, position) #this function turns df_raw into df_calculated
    ###print(df_calculated)
    ###print("----------------")

    winrate_dict = winrate.create_df_winrate(df_raw)
    #print(winrate_dict) # more important info is printing the df_isVictory in winrate.py

    df_player_calculations, parameters_dict = calculations.player_calculations(df_calculated, steam_id, minute, isOnMyTeam, number_of_matches_to_parse, position, winrate_dict)
    print(parameters_dict)
    print(df_player_calculations)

    #df_player_calculations.to_excel("output.xlsx", sheet_name="Sheet1", index=False)

    # Database interfacing functions
    flat_df = database_handler.table_player_calculations_staging(df_player_calculations, parameters_dict, engine)
    database_handler.join_tables(flat_df, engine)

    # Creates plots 
    dict_of_plts = calculations.player_graphs(df_calculated, position) #changes paramaters to get different members of your team ("POSITION_2", isOnMyTeam=False for enemy mid)
    print(f"Number of matches parsed: {(df_calculated.shape[0])/10}")

    # Creates excel sheet
    make_all_excel_sheets(df_raw, df_player_calculations, dict_of_plts, steam_id, position, minute, number_of_matches_to_parse, isOnMyTeam)

    return df_player_calculations, parameters_dict


if __name__ == "__main__":
    main_script(steam_id, position, isOnMyTeam, minute, skip_interval, number_of_matches_to_parse)


